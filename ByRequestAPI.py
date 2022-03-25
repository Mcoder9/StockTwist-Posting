import requests,json,threading,random,time,names,os
from bs4 import BeautifulSoup
from colorama import Fore
from openpyxl import Workbook,load_workbook
##############################
if os.path.exists('Accounts.xlsx'):
    wb = load_workbook('Accounts.xlsx')
else:
    wb = Workbook()
ws = wb['Sheet']
Proxy_username = 'mgcoder5'
Proxy_password = 'OnceMoretry!27'
proxy = f'http://{Proxy_username}:{Proxy_password}@gate.smartproxy.com:7000'
f = open('avatar.jpeg','rb')
File_data = f.read()
f.close()
def make_post(body,token):
    url = "https://api.stocktwits.com/api/2/messages/create.json?cc_to_twitter=0"

    payload={'body': body}
    headers = {
    'authorization': f'OAuth {token}',
    "origin": "https://stocktwits.com",
    "referer": "https://stocktwits.com/",
    "accept": "application/json",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.74 Safari/537.36"
    }

    response = requests.request("POST", url, headers=headers, data=payload, proxies={'http': proxy, 'https': proxy})
    return response
def makesoup(url):
    while True:
        try:
            res = requests.get(url)
            break
        except:
            pass
    soup = BeautifulSoup(res.text, 'lxml')
    content = soup.get_text()
    return content
def start_posting(token):
    Tickers = ["AAPL", "TSLA", "AMZN", "FB", "NFLX", "SPY", "AMD", "MSFT", "AMC", "NIO", "BABA", "NVDA", "TWTR", "DIS", "GME", "GOOG", "SNDL", "SNAP", "PLTR", "LCID", "OCGN", "BA", "PLUG", "SQ", "F", "GE", "BB", "FCEL", "SPCE", "ZOM", "TLRY", "MU", "WKHS", "BAC", "GOOGL", "QQQ", "GEVO", "NAKD", "ACB", "INTC", "IDEX", "ROKU", "BNGO", "MARA", "NOK", "PYPL", "GNUS", "WMT", "RIOT", "AAL", "IBIO", "INO", "SBUX", "DKNG", "CEI", "SHOP", "MRNA", "T", "NKE", "CTRM", "UBER", "GPRO", "NKLA", "CLOV", "SOS", "MVIS", "NVAX", "CGC", "PFE", "BYND", "COIN", "VXRT", "TOPS", "V", "JAGX", "XSPA", "GILD", "CCL", "FSR", "WISH", "XOM", "NNDM", "SRNE", "JPM", "DAL", "ATOS", "UVXY", "CRON", "VISL", "TNXP"]
    for i,x in enumerate(Tickers,start=1):
        # With URL
        url = f'https://www.tslatrading.com/calls/index.php?t={x}'
        soup = makesoup(url)
        # without URL
        output = ""
        for text in soup.split('\n')[:-1]:
            k = text.strip()
            output+=k + '\n'
        output = output.strip()
        topost = ''
        if start_lposting.lower() == 'y' or i>14:
            topost = soup.strip()
        else:
            topost = output
        print(topost)
        response = make_post(topost,token)
        if response.status_code==401:
            print(Fore.RED + "Blocked" +Fore.RESET)
            return
        elif response.status_code == 200:
            print(Fore.GREEN + "Successfully Posted..." + Fore.RESET)
            return
        else:
            print(Fore.RED + 'Something go wrong' + Fore.RESET)

def update_avatar(token):
    header = {"authorization":f"OAuth {token}"}
    files = {'avatar':File_data}
    while True:
        try:
            r = requests.post('https://api.stocktwits.com/api/2/account/update_avatar.json',files=files,headers=header, proxies={'http': proxy, 'https': proxy})
            break
        except:
            pass
    if r.status_code == 200:
        print(Fore.YELLOW + 'Profile Pic Updated.' + Fore.RESET)
    else:
        print(Fore.RED + 'Error In Profile Updation.' + Fore.RESET)

def gen_email(username,name):
    
    def Check_msg(email,timestemp):
        startTime = int(time.time())
        while True:
            key_pay = {
            "email":email,
            "timestamp":timestemp,
            "server":"server-1",
            "type":"alias"
            }
            while True:
                try:
                    r = requests.post('https://smailpro.com/app/key',data=key_pay, proxies={'http': proxy, 'https': proxy})
                    break
                except:
                    pass
            key = json.loads(r.text)['items']
            url = f"https://public-sonjj.p.rapidapi.com/email/gm/check?key={key}&rapidapi-key=f871a22852mshc3ccc49e34af1e8p126682jsn734696f1f081&email={email}&timestamp={timestemp}&server=server-1&type=alias"
            while True:
                try:
                    r = requests.get(url, proxies={'http': proxy, 'https': proxy})
                    break
                except:
                    pass
            try:
                items = json.loads(r.text)['items']
            except:
                return
            if len(items)>0:
                return items[0]
            change = int(time.time())-startTime
            if change>90:
                return 0
    def Parse_sms(item):
        try:
            payload = {
                "email":item['textTo'],
                "message_id":item['mid'],
                "server":"server-1",
                "type":"alias"
            }
        except:
            return 0
        r = requests.post('https://smailpro.com/app/key',data=payload, proxies={'http': proxy, 'https': proxy})
        key = json.loads(r.text)['items']
        url = "https://public-sonjj.p.rapidapi.com/email/gm/read"
        read_payload = {
            "key": key,
            "rapidapi-key": "f871a22852mshc3ccc49e34af1e8p126682jsn734696f1f081",
            "email": item['textTo'],
            "message_id": item['mid'],
            "server": "server-1",
            "type": "alias",
        }
        r = requests.get('https://public-sonjj.p.rapidapi.com/email/gm/read',params=read_payload, proxies={'http': proxy, 'https': proxy})
        body = json.loads(r.text)['items']['body']
        return body
    def html_parse(source):
        soup = BeautifulSoup(str(source),'html.parser')
        ar = soup.findAll('a')
        for a in ar:
            href = a['href']
            if str(href).find("token")>-1:
                token = str(href).split('token=')[1]
                return token
    key_payload={"username":"random",
    "domain":"gmail.com",
    "server":"server-1",
    "type":"alias"
    }
    while True:
        try:    
            r = requests.post('https://smailpro.com/app/key',data=key_payload, proxies={'http': proxy, 'https': proxy})
            if r.status_code != 200:
                return
            break
        except:
            pass
    if r.status_code == 200:
        key = json.loads(r.text)['items']
        email_params = {
            "key": key,
            "rapidapi-key": "f871a22852mshc3ccc49e34af1e8p126682jsn734696f1f081",
            "username": "random",
            "domain": "gmail.com",
            "server": "server-1",
            "type": "alias",
        }
        r = requests.get('https://public-sonjj.p.rapidapi.com/email/gm/get',params=email_params, proxies={'http': proxy, 'https': proxy})
        email = json.loads(r.text)['items']['email']
        print(f"Email: {email}")
        time_stemp = json.loads(r.text)['items']['timestamp']
        payload={
        "newsletter": 1,
        "captchaToken": "03AGdBq27x-b9Krfm96uZo56Aaoqj-ckmf_CytoWKMDglIlQP11eUo6XgCerLi-cdMuXiVfCKyZhCkCOyT6ypE5eMJIc6UsQ6uxj77LzAr9ZqIlCT91K2JkMdMV8IEMw36F7JMH7MLdLxERFjfnSM4RH_56Udrum4oHqg4hpJhZrxORBDa5AMHoPEvO_gygLhiRO2VN6Fn-AkaLRDiQFUpZA5PDkcm9IbwafMAUUmnbDi6jigRlOZcXRcog08ScJP6Z1jd3tKOPLJmmIRet3fk5ZkMbfhCQb2tmpgOv2uzJRw5ZAlZ3HB_ug322b74dgVW0ZbuYm8DP2mgwF69zk42qWmD5V7wryOYp--KARW_71sMOalHDmq19sS9SJPPi5iGzZ8k-dOIoYPLdZ5ymGfQ6FZlyjP4kamE7k2RSR35kvRxUVy5ebmLyXKu8-xs3SuqELiluZ_ye0lH",
        "name": name,
        "email": email,
        "login": username,
        "password": "Pakistan99"
        }
        r = requests.post('https://stocktwits.com/api/register',data = payload, proxies={'http': proxy, 'https': proxy})
        if r.status_code != 200:
            print(Fore.RED + f'{email}: Not Register' + Fore.RESET)
        if r.status_code==200:
            try:
                mainToken = json.loads(r.text)['token']
            except:
                print(Fore.RED + f'{email}: Not Register' + Fore.RESET)
                return
            print(Fore.YELLOW + f"Token Generated: {mainToken}" + Fore.RESET)
            header = {
                "authorization":f"OAuth {mainToken}"
            }
            t = threading.Thread(target=update_avatar,args=(mainToken,))
            t.start()
            for i in range(2):
                r = requests.get('https://api.stocktwits.com/api/2/account/verify_resend',headers=header, proxies={'http': proxy, 'https': proxy})
                if r.status_code == 200:
                    print("[INFO] Verification Sent")
                else:
                    print(r.text)
                    print(Fore.RED + "[Err] Verification Not Sent" + Fore.RESET)
                    return 0
                print("[INFO] Waiting for Reply")

                item = Check_msg(email,time_stemp)
                if item == 0:
                    print(Fore.RED + "Message Not Recieve")
                else:
                    break
            body = Parse_sms(item)
            if body == 0:
                print('Not registored')
                return
            vtoken = html_parse(body)
            token_payload = {"token":f"{vtoken}"}
            r = requests.post('https://api.stocktwits.com/api/2/account/verify_identity.json',headers=header,data=token_payload, proxies={'http': proxy, 'https': proxy})
            if r.status_code == 200 :
                print(Fore.GREEN + f"Token Verified: {mainToken}" + Fore.RESET)
                ws.append([username,email,"Pakistan99",mainToken])
                wb.save('Accounts.xlsx')
                start_posting(mainToken)
            else:
                print(Fore.RED + f"Token Not Verified: {mainToken}" + Fore.RESET)
                t.join()
                return
            t.join()
            
            # accountsf.write(f'"{username}","{email}","Pakistan99","{mainToken}"\n')

def main_fun():
    name = names.get_full_name()
    username = name.replace(' ', str(random.randint(1,500))).replace('.','_')
    gen_email(f"{username}{i}",name)
if __name__ =="__main__":
    x = int(input(Fore.LIGHTBLUE_EX+'No of Accounts? ' + Fore.RESET))
    while True:
        start_lposting = input('Start Posting With Links(Y/N)? ')
        start_lposting = start_lposting.lower()
        if start_lposting !='y' and start_lposting !='n':
            print("Invalid Choice Try Again...")
        else:
            break
    for i in range(x):
        main_fun()